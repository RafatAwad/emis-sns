from datetime import datetime,timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

from django.shortcuts import render, redirect
from .forms import StudentForm, StudentAddressForm,ParentAddressForm, StudentDocumentForm, StudentSchoolForm, StudentParentForm,StudentAddressForm2,ParentAddressForm
from .filters import StudentFilter,SchoolFilter
from .models import StudentInfo,School,SchoolLevel
from address.models import Location, District, Governorate,SubDistrict,Village,SubVillage
from django.contrib import messages
from django.core.paginator import Paginator
from django.utils.translation import gettext_lazy as _
from django.contrib.auth.decorators import login_required
from account.decorators import is_not_admin, user_authenticated
import random
from django.utils.text import Truncator
from datetime import date
current_year = str(date.today().year)
stu_year= current_year[2]+current_year[3]

def n_len_rand(len_, floor=1):
    top = 10**len_
    if floor > top:
        raise ValueError(f"Floor '{floor}' must be less than requested top '{top}'")
    return f'{random.randrange(floor, top):0{len_}}'
 
@login_required
def student_list(request):
    context= {}
    if request.user.role.id == 3 :
        student = StudentInfo.objects.filter(school__school_address__district = request.user.location.district).order_by('-added_date')
        context.update({'gov': request.user.location.district.id})
    elif request.user.role.id == 2:
        student = StudentInfo.objects.filter(school__school_address__governorate = request.user.location.governorate).order_by('-added_date')
        districts = District.objects.filter(governorate = request.user.location.governorate)
        context.update({'districts': districts,'gov': request.user.location.governorate.id})
    else :
        student = StudentInfo.objects.all().order_by('-added_date')
        context.update({'gov': 1})

    tableFilter = StudentFilter(request.GET, queryset=student)
    student = tableFilter.qs
    paginator = Paginator(student, 20) 
    page_number = request.GET.get('page')
    students_with_pag = paginator.get_page(page_number)
    context.update({'items': students_with_pag,
               'tableFilter': tableFilter,
                'url' : "student-list2"
               })
    return render(request, 'student/student-list.html', context)
@login_required
def student_list2(request):
    context= {}
    if request.user.role.id == 3 :
        student = StudentInfo.objects.filter(school__school_address__district = request.user.location.district).order_by('-added_date')
    elif request.user.role.id == 2:
        student = StudentInfo.objects.filter(school__school_address__governorate = request.user.location.governorate).order_by('-added_date')
        districts = District.objects.filter(governorate = request.user.location.governorate)
        context.update({'districts': districts})
    else :
        student = StudentInfo.objects.all().order_by('-added_date')
    tableFilter = StudentFilter(request.GET, queryset=student)
    student = tableFilter.qs
    paginator = Paginator(student, 20) 
    page_number = request.GET.get('page')
    students_with_pag = paginator.get_page(page_number)
    context.update({'items': students_with_pag,
               'tableFilter': tableFilter,
                'url' : "student-list2"
               })
    return render(request, 'partial/students.html', context)

#schools list  for all roles
@login_required
def schools_list(request):
    levels = SchoolLevel.objects.all()
    context = {}
    if request.user.role.id == 3 :
        school = School.objects.filter(school_address__district = request.user.location.district).order_by('school_name')

    elif request.user.role.id == 2:
        school = School.objects.filter(school_address__governorate = request.user.location.governorate).order_by('school_address__district','school_name')
        districts = District.objects.filter(governorate = request.user.location.governorate)
        context.update({'districts': districts})
    else:
        school = School.objects.order_by('school_address__governorate')
    tableFilter = SchoolFilter(request.GET, queryset=school)
    school = tableFilter.qs
    paginator = Paginator(school, 20) 
    page_number = request.GET.get('page')
    schools_with_pag = paginator.get_page(page_number)
    context.update({'items': schools_with_pag,
                'levels': levels,
                'tableFilter': tableFilter,
                'url' : "schools-list2"
               })
    return render(request, 'student/schools-list.html', context)

@login_required
def schools_list2(request):
    levels = SchoolLevel.objects.all()
    context = {}
    if request.user.role.id == 3 :
        school = School.objects.filter(school_address__district = request.user.location.district).order_by('school_name')

    elif request.user.role.id == 2:
        school = School.objects.filter(school_address__governorate = request.user.location.governorate).order_by('school_address__district','school_name')
        districts = District.objects.filter(governorate = request.user.location.governorate)
        context.update({'districts': districts})
    else:
        school = School.objects.order_by('school_address__governorate')
    tableFilter = SchoolFilter(request.GET, queryset=school)
    school = tableFilter.qs
    paginator = Paginator(school, 20) 
    page_number = request.GET.get('page')
    schools_with_pag = paginator.get_page(page_number)
    context.update({'items': schools_with_pag,
                'levels': levels,
                'tableFilter': tableFilter,
                 'url' : "schools-list2"
               })
    return render(request, 'partial/schools.html', context)

@login_required
def add_student(request, school_code):
    school = School.objects.get(school_code = school_code)
    student_form = StudentForm(request.POST or None, request.FILES or None)
    student_parent_form = StudentParentForm(request.POST or None, request.FILES or None)
    parent_address_form = ParentAddressForm(request.POST or None)
    student_documents_form = StudentDocumentForm(request.POST or None, request.FILES or None)
    if request.method == 'POST':
        if student_form.is_valid() and student_parent_form.is_valid() and parent_address_form and student_documents_form.is_valid():
            if request.POST['school_code']:
                student_school = School.objects.get(school_code = request.POST['school_code'])
                student_address_form = Location.objects.get(id = request.POST['address'])
            #rand_num = n_len_rand(4)
            student_count = StudentInfo.objects.filter(school = request.POST.get('school')).count()
            student_number = str(student_count+1).zfill(4)
            s1 = student_school
            emis_id = str(s1.school_code.zfill(5))+"_"+ str(stu_year)+"_" + str(student_number)
            s2 = student_address_form
            s3 = student_parent_form.save(commit=False)
            if request.POST.get('district'):
                 s3.parent_address = Location.objects.get(id = request.POST.get('district'))
            elif request.POST.get('governorate_p'):
                 s3.parent_address = Location.objects.get(id = request.POST.get('governorate'))            
            
            s3.save()
            s4 = student_documents_form.save()
            student_info = student_form.save(commit=False)
            student_info.school = s1
            student_info.address = s2
            student_info.parent = s3
            student_info.documents = s4
            student_info.emis_id = emis_id
            student_info.added_by = request.user
            student_info.save()
            messages.success(request,"تم إضافة الطالب بنجاح")
            return redirect('student-list')
    context = {
        'student_form': student_form,
        'school_info': school,
        'parent_form': student_parent_form,
        'parent_address_form': parent_address_form,
        'documents_form': student_documents_form
    }
    return render(request, 'student/add-student.html', context)

#@login_required
def student_profile(request, emis_id):
    student = StudentInfo.objects.get(emis_id=emis_id)
    context = {
        'student': student
    }
    if  user_authenticated :
        return render(request, 'student/stdform2.html', context)
    else:
        return render(request, 'student/stdform.html', context)



@login_required
@is_not_admin(redirect_url = 'schools-list')
def student_registration(request):
    student_form = StudentForm(request.POST or None, request.FILES or None)
    student_address_form = StudentAddressForm2(request.POST or None)
    student_parent_form = StudentParentForm(request.POST or None, request.FILES or None)
    parent_address_form = ParentAddressForm(request.POST or None)
    student_documents_form = StudentDocumentForm(request.POST or None, request.FILES or None)

    if request.method == 'POST':
        if student_form.is_valid() and student_address_form.is_valid() and student_parent_form.is_valid() and parent_address_form and student_documents_form.is_valid():
            if request.POST['sub_village']:
                address =Location.objects.get(id = request.POST['sub_village'])
            elif request.POST['village']:
                address =Location.objects.get(id = request.POST['village'])
            elif request.POST['sub_district']:
                address =Location.objects.get(id = request.POST['sub_district'])
            elif request.POST['district'] :
                address =Location.objects.get(id = request.POST['district'])
            elif request.POST['governorate'] :
                address =Location.objects.get(id = request.POST['governorate'])
            else:
                address = None
            if request.POST['school']:
                s1 = School.objects.get(id = request.POST['school'])
            #rand_num = n_len_rand(4)
            student_count = StudentInfo.objects.filter(school = request.POST['school']).count()
            student_number = str(student_count+1).zfill(4)
            emis_id = str(s1.school_code.zfill(5))+"_"+ str(stu_year)+"_" + str(student_number)
            s2 = address
            s3 = student_parent_form.save(commit=False)
            if request.POST.get('district'):
                 s3.parent_address = Location.objects.get(id = request.POST.get('district'))
            elif request.POST.get('governorate_p'):
                 s3.parent_address = Location.objects.get(id = request.POST.get('governorate'))            
            
            s3.save()
            s4 = student_documents_form.save()
            student_info = student_form.save(commit=False)
            student_info.emis_id = emis_id
            student_info.school = s1
            student_info.address = s2
            student_info.parent = s3
            student_info.documents = s4
            student_info.save()
            messages.success(request,"تم إضافة الطالب بنجاح")
            return redirect('student-list')
    context = {
        'student_form': student_form,
        'address_form': student_address_form,
        'parent_form': student_parent_form,
        'parent_address_form': parent_address_form,
        'documents_form': student_documents_form
    }
    return render(request, 'student/student-registration.html', context)


@login_required
def student_edit(request, emis_id):
    main = {}
    context = {}
    student = StudentInfo.objects.get(emis_id=emis_id)
    main.update({
        'name': student.name,
        'address': student.address,
        'school': student.school,
    })
    student_form = StudentForm(instance=student)
    student_school_form = StudentSchoolForm(instance=student.school)
    student_address_form = StudentAddressForm(instance=student.address)
    student_Parent_form = StudentParentForm(instance=student.parent)
    Parent_address_form = ParentAddressForm(instance=student.parent.parent_address)
    student_documents_form = StudentDocumentForm(instance=student.documents)
    if request.method == 'POST':
        student_form = StudentForm(request.POST, request.FILES, instance=student)
        student_Parent_form = StudentParentForm(request.POST, request.FILES, instance=student.parent)
        student_documents_form = StudentDocumentForm(request.POST, request.FILES, instance=student.documents)
        if student_form.is_valid() and student_Parent_form.is_valid() and Parent_address_form and student_documents_form.is_valid():
            s3 = student_Parent_form.save(commit=False)
            if request.POST.get('district'):
                 s3.parent_address = Location.objects.get(id = request.POST.get('district'))
            elif request.POST.get('governorate_p'):
                 s3.parent_address = Location.objects.get(id = request.POST.get('governorate'))

            s3.save()
            s4 = student_documents_form.save()
            student_info = student_form.save(commit=False)
            if request.POST.get('school') :
                school = School.objects.get(id = request.POST.get('school'))
                location = Location.objects.get(id = school.school_address.id)
                student_address_form = StudentAddressForm(request.POST,instance=location)
                s1 = school
                student_info.school = s1
                s2 = location
                student_info.address = s2
                
            student_info.parent = s3
            student_info.documents = s4
            student_info.save()
            messages.success(request, "تمت عملية التعديل" )
            return redirect('student-list')
        messages.error(request, "لم تتم عملية التعديل")

    context.update({
        'student_form': student_form,
        'school_form': student_school_form,
        'address_form': student_address_form,
        'parent_form': student_Parent_form,
        'parent_address_form': Parent_address_form,
        'documents_form': student_documents_form,
        'main' : main
    })
    return render(request, 'student/student-edit.html', context)


@login_required
def check_edit(request, emis_id):
    main = {}
    context = {}
    student = StudentInfo.objects.get(emis_id=emis_id)
    main.update({
        'name': student.name,
        'address': student.address,
        'school': student.school,
    })
    student_form = StudentForm(instance=student)
    student_school_form = StudentSchoolForm(instance=student.school)
    student_address_form = StudentAddressForm(instance=student.address)
    student_Parent_form = StudentParentForm(instance=student.parent)
    Parent_address_form = ParentAddressForm(instance=student.parent.parent_address)
    student_documents_form = StudentDocumentForm(instance=student.documents)
    if request.method == 'POST':
        student_form = StudentForm(request.POST, request.FILES, instance=student)
        student_Parent_form = StudentParentForm(request.POST, request.FILES, instance=student.parent)
        student_documents_form = StudentDocumentForm(request.POST, request.FILES, instance=student.documents)
        if student_form.is_valid() and student_Parent_form.is_valid() and Parent_address_form and student_documents_form.is_valid():
            s3 = student_Parent_form.save(commit=False)
            if request.POST.get('district'):
                 s3.parent_address = Location.objects.get(id = request.POST.get('district'))
            elif request.POST.get('governorate_p'):
                 s3.parent_address = Location.objects.get(id = request.POST.get('governorate'))

            s3.save()
            s4 = student_documents_form.save()
            student_info = student_form.save(commit=False)
            if request.POST.get('school') :
                school = School.objects.get(id = request.POST.get('school'))
                location = Location.objects.get(id = school.school_address.id)
                student_address_form = StudentAddressForm(request.POST,instance=location)
                s1 = school
                student_info.school = s1
                s2 = location
                student_info.address = s2

            student_info.parent = s3
            student_info.documents = s4
            student_info.checked = True
            student_info.checked_by = request.user
            student_info.checked_date = datetime.now()
            student_info.note = request.POST.get('notes')
            student_info.save()
            messages.success(request, "تمت عملية التحقق" )
            return redirect('student-list')
        messages.error(request, "لم تتم عملية التحقق")

    context.update({
        'student_form': student_form,
        'school_form': student_school_form,
        'address_form': student_address_form,
        'parent_form': student_Parent_form,
        'parent_address_form': Parent_address_form,
        'documents_form': student_documents_form,
        'main' : main
    })
    return render(request, 'student/check-edit.html', context)


@login_required
def sch_edit(request, school_code):
    school = School.objects.get(school_code =school_code)
    school_form = StudentSchoolForm(instance=school)
    if request.method == 'POST':
        school_form = StudentSchoolForm(request.POST, request.FILES, instance=school)
        if school_form.is_valid() :
            school_form.save()

            messages.success(request, "تمت عملية التعديل" )
            return redirect('schools-list')
        messages.error(request, "لم تتم عملية التعديل")

    context = {
        'school_form': school_form,
    }
    return render(request, 'student/school-edit.html', context)

@login_required
def student_delete(request, emis_id):
    student = StudentInfo.objects.get(emis_id=emis_id)
    student.delete()
    return redirect('student-list')


@login_required
def sch_delete(request, id):
    school = School.objects.get(id=id)
    school.delete()
    return redirect('schools-list')

@login_required
def print_form(request):
    return render(request, 'student/student-form.html')

@login_required
def sch_profile(request, id):
    school = School.objects.get(id=id)
    context = {
        'school': school
    }
    return render(request, 'student/school-profile.html', context)
@login_required
def check_now(request, emis_id):
    student = StudentInfo.objects.get(emis_id =emis_id)
    student.checked = True
    student.checked_by = request.user
    student.checked_date = datetime.now()
    student.save()
    messages.success(request, _('student checked successfully'))
    return redirect('student-list')

@login_required
def recheck(request, emis_id):
    student = StudentInfo.objects.get(emis_id =emis_id)
    student.checked = False
    student.checked_by = None
    student.checked_date = None
    student.note = None
    student.save()
    messages.success(request, _('student now unchecked '))
    return redirect('student-list')

@login_required
def gov_schools(request, id):
    context = {}
    if request.user.role.id == 1 :
        schools = School.objects.filter(school_address__governorate = id).order_by('school_name')
        address = Governorate.objects.filter(id = id).first()
        districts = District.objects.filter(governorate = id)
        context.update({
            "districts" : districts,
            "gov" : id,
        })
    else :
        schools = School.objects.filter(school_address__district = id).order_by('school_name')
        address = District.objects.filter(id = id).first()
        context.update({
            "dist" : id,
            "gov" : address.governorate.id,
        })

    tableFilter = SchoolFilter(request.GET, queryset=schools)
    schools = tableFilter.qs
    paginator = Paginator(schools, 20) 
    page_number = request.GET.get('page')
    schools_with_pag = paginator.get_page(page_number)
   
    context.update({
        'items': schools_with_pag,
         'tableFilter': tableFilter,
         'address': address,
         'url':  'schools-list2',

    })
    return render(request, 'student/schools-list2.html', context)


def export_stu(request, gov):
    """
    Downloads all movies as Excel file with a worksheet for each movie category
    """
    if request.user.role.id == 1 :
        students = StudentInfo.objects.all()
        title = "طلاب جميع المحافظات"
    elif request.user.role.id == 2 :
        students = StudentInfo.objects.filter(school__school_address__governorate = gov)
        title =   request.user.location.name
    else:
        students = StudentInfo.objects.filter(school__school_address__district = gov)
        title =   request.user.location.name
        
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-students.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()
    
    header_font = Font(name='Calibri', bold=True,size=14)
    body_font = Font(name='arial', bold=True,size=12)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='center',
        horizontal='center',
        wrap_text=True
    )
    fill = PatternFill(
    start_color='c7dee1',
    end_color='c7dee1',
    fill_type='solid',
    )
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = ' طلاب' + title
    worksheet.font = Font(name='Calibri', bold=True, size=20)
    worksheet.sheet_properties.tabColor = 'c7dee1'

    # Define the titles for columns
    columns = [
        ('ID',6),
        ('اسم الطالب',24),
        ('الرقم الوطني',20),
        ('المحافظة',15),
        ('المديرية',15),
        ('المدرسة',20),
        ('الصف',10),
        ('نوع الإعاقة',10),
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, (column_title, column_width) in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.border = border_bottom
        cell.alignment = centered_alignment
        cell.fill = fill
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = column_width
    # Iterate through all movies
    for student in students:
        row_num += 1
        
        # Define the data for each cell in the row 
        row = [
            student.pk,
            student.name,
            student.emis_id,
            student.address.governorate.name,
            student.address.district.name,
            student.school.school_name,
            student.grade.name,
            student.spn_type.name,
        ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
            cell.font = body_font


    worksheet.freeze_panes = worksheet['A1']
    workbook.save(response)

    return response